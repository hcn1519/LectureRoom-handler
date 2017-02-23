import openpyxl
import re
import numpy as np
from collections import OrderedDict


def searchRoom(lectureRoom, roomName):
    x=1
    for x in range(len(lectureRoom[x])):
        for y in range(len(lectureRoom)):
            if(lectureRoom[y][3] == roomName):
                return lectureRoom[y][0]

wb = openpyxl.load_workbook('/Users/changnam/Desktop/lectureSeed.xlsx')

sheet1 = wb.get_sheet_by_name('Sheet1')

lectureRoom = [[0 for x in range(4)] for y in range(sheet1.max_row)]

lectureTime = [[0 for x in range(5)] for y in range(sheet1.max_row*2)]
lectureTime[0][0] = "id"
lectureTime[0][1] = "lectureRoom_id"
lectureTime[0][2] = "dayOftheWeek"
lectureTime[0][3] = "from"
lectureTime[0][4] = "to"


for i in range(1, sheet1.max_row):
    flag1 = re.match(r"(\w+)[,](\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))
    flag2 = re.match(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[/]\s(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))
    flag3 = re.match(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))

    if(flag1): # "wed,fri 09:00~10:15 [J602]", # tue,thu 15:00~16:15 [D209A]
        form1 = re.split(r"(\w+)[,](\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))
        form1 = [x for x in form1 if x != '']
        lectureRoom[i][1] = form1[4]
        lectureRoom[i][2] = form1[5]
        lectureRoom[i][3] = form1[4]+form1[5]
    elif(flag2): # "mon 10:30~11:45 / wed 12:00~13:15 [J313]", mon 10:30~11:45 / wed 12:00~13:15 [RA406A]
        form2 = re.split(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[/]\s(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))
        form2 = [x for x in form2 if x != '']
        lectureRoom[i][1] = form2[6]
        lectureRoom[i][2] = form2[7]
        lectureRoom[i][3] = form2[6]+form2[7]
    elif(flag3): # thu 18:00~20:45 [J114], thu 18:00~20:45 [J114A]
        form3 = re.split(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=i+1, column=2).value.encode("utf-8"))
        form3 = [x for x in form3 if x != '']
        lectureRoom[i][1] = form3[3]
        lectureRoom[i][2] = form3[4]
        lectureRoom[i][3] = form3[3]+form3[4]


data = np.array(lectureRoom)

ncols = data.shape[1]
dtype = data.dtype.descr * ncols
struct = data.view(dtype)

lectureRoom = np.unique(struct)
lectureRoom = lectureRoom.view(data.dtype).reshape(-1, ncols)

lectureRoom = np.insert(lectureRoom, 0, np.array(("id", "building", "room no", "name")), 0)
lectureRoom = np.delete(lectureRoom, 1, 0)

id=1
for id in range(len(lectureRoom[id])):
    for num in range(len(lectureRoom)):
        if(num ==0): continue
        lectureRoom[num][0] = num


for m in range(1, sheet1.max_row):
    flag1 = re.match(r"(\w+)[,](\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
    flag2 = re.match(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[/]\s(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
    flag3 = re.match(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
    if(flag1): # "wed,fri 09:00~10:15 [J602]", # tue,thu 15:00~16:15 [D209A]
        form1 = re.split(r"(\w+)[,](\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
        form1 = [x for x in form1 if x != '']
        a=1
        while(a<3):
            lectureTime[m*2-(a%2)][0] = m*2-(a%2)
            lectureTime[m*2-(a%2)][1] = searchRoom(lectureRoom, form1[4]+ form1[5])
            if(a%2 == 1):
                lectureTime[m*2-(a%2)][2] = form1[0]
            elif(a%2 == 0):
                lectureTime[m*2-(a%2)][2] = form1[1]
            lectureTime[m*2-(a%2)][3] = form1[2]
            lectureTime[m*2-(a%2)][4] = form1[3]
            a=a+1
    elif(flag2): # "mon 10:30~11:45 / wed 12:00~13:15 [J313]", mon 10:30~11:45 / wed 12:00~13:15 [RA406A]
        form2 = re.split(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[/]\s(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
        form2 = [x for x in form2 if x != '']
        a=1
        while(a<3):
            lectureTime[m*2-(a%2)][0] = m*2-(a%2)
            lectureTime[m*2-(a%2)][1] = searchRoom(lectureRoom, form2[6]+ form2[7])

            #dayOftheWeek, from ~ to
            if(a%2 == 1):
                lectureTime[m*2-(a%2)][2] = form2[0]
                lectureTime[m*2-(a%2)][3] = form2[1]
                lectureTime[m*2-(a%2)][4] = form2[2]
            elif(a%2 == 0):
                lectureTime[m*2-(a%2)][2] = form2[3]
                lectureTime[m*2-(a%2)][3] = form2[4]
                lectureTime[m*2-(a%2)][4] = form2[5]
            a=a+1
    elif(flag3): # thu 18:00~20:45 [J114], thu 18:00~20:45 [J114A]
        form3 = re.split(r"(\w+)\s(\w+[:]\w+)[~](\w+[:]\w+)\s[\[]([a-zA-Z]+)(\w+)[\]]", sheet1.cell(row=m+1, column=2).value.encode("utf-8"))
        form3 = [x for x in form3 if x != '']
        # print form3
        a=1
        lectureTime[m*2-(a%2)][0] = m*2-(a%2)
        lectureTime[m*2-(a%2)][1] = searchRoom(lectureRoom, form3[3]+ form3[4])
        lectureTime[m*2-(a%2)][2] = form3[0]
        lectureTime[m*2-(a%2)][3] = form3[1]
        lectureTime[m*2-(a%2)][4] = form3[2]


# delete 0 value
lectureTime = np.array(lectureTime)
for i in range(len(lectureTime[i])):
    for j in range(len(lectureTime)):
        if(j == 1900): break
        if(lectureTime[j][1] == 0 or lectureTime[j][1] == '0'):
            lectureTime = np.delete(lectureTime, j, 0)
h=1
for h in range(len(lectureTime[h])):
    for u in range(len(lectureTime)):
        if(u != 0):
            lectureTime[u][0] = u

lectureTime = np.delete(lectureTime, 1900, 0)
print lectureTime

workb = openpyxl.Workbook()
ws = workb.create_sheet(title="LectureRoom")

i=0
for i in range(0, len(lectureRoom[i])):
    for j in range(0, len(lectureRoom)):
       ws.cell(column=i+1, row=j+1, value=lectureRoom[j][i])

workb.save(filename='LectureRoom.xlsx')

workb2 = openpyxl.Workbook()
ws2 = workb2.create_sheet(title="LectureTime")

q=0
for q in range(0, len(lectureTime[q])):
    for w in range(0, len(lectureTime)):
        ws2.cell(column=q+1, row=w+1, value=lectureTime[w][q])

workb2.save(filename='LectureTime.xlsx')

# mon 10:30~11:45 / wed 12:00~13:15 [J313]
# wed,fri 09:00~10:15 [J602]
# tue,thu 15:00~16:15 [D209A]
# thu 18:00~20:45 [J114]
# thu 18:00~20:45 [J114A]
# mon 10:30~11:45 / wed 12:00~13:15 [RA406A]
